import os
import re
import sys
import math
import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as soup
from datetime import datetime, timedelta
from itertools import permutations
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import csv
from pathlib import Path
import time
import warnings

# Supprimer les avertissements pandas FutureWarning
warnings.filterwarnings('ignore', category=FutureWarning)

####################### CONFIGURATION #######################
# Chemins locaux
CSV_FILENAME = './data/joueurs_ATP_WTA.csv'  # Chemin local pour le fichier joueurs
GECKODRIVER_PATH = './geckodriver'  # Chemin vers Geckodriver
DATA_FOLDER = './data'  # Dossier pour les fichiers générés
#############################################################


############# ANNEXE

audit = False# True False # Mettre à False pour désactiver l'audit
audit1 = False# True False

def listing(datas):
    if datas:
        for data in datas:
            print(len(data),data)

def extract_source(url):
    """Extraire le HTML d'une URL."""
    try:
        with requests.Session() as session:
            session.headers.update({
                "User-Agent": "Mozilla/5.0",
                "Accept-Encoding": "gzip, deflate"
            })
            response = session.get(url, timeout=10)
            response.raise_for_status()
            return soup(response.content, 'html.parser')
    except requests.exceptions.RequestException as e:
        print(f"Request error: {e}")
        return None

def calculer_ratio(data, separateur):
    """Calculer le ratio en pourcentage à partir des données."""
    if data == '-' and separateur in ['-', '/']:
        data = '0-0' if separateur == '-' else '0/0'
    if data == 'None' or data.split(separateur)[0] == 'None':
        return 0, 0
    numerator, denominator = map(int, data.split(separateur))
    total = numerator + denominator
    return (numerator * 100 / total if total else 0), (denominator * 100 / total if total else 0)

def paires(liste):
    """Créer des paires à partir d'une liste."""
    if len(liste) % 2 != 0:
        print(f"Avertissement: La liste contient un nombre impair d'éléments ({len(liste)}). Ajout d'un élément par défaut.")
        # Ajouter un élément par défaut pour équilibrer la liste
        liste = liste + ['']
    return [liste[i:i + 2] for i in range(0, len(liste), 2)]

def Lien_Google(tabF):
    """Générer des liens Google pour les matchs."""
    return [['https://www.google.com/search?client=firefox-b-d&q=Tennis+' + match[0].replace(' ', '+')] for match in tabF]

def name_onglet(url):
    # Utilisation d'une expression régulière pour extraire le texte entre "https://www.tennisexplorer.com/" et le "/" suivant
    match = re.search(r'https://www.tennisexplorer.com/([^/]+)/', url)

    if match:
        # Obtention du texte extrait
        extracted_text = match.group(1)
        return extracted_text
    else:
        # Retourner None si l'URL ne correspond pas au modèle attendu
        return None

def safe_ceil(value):
    try:
        return str(math.ceil(float(value)))
    except ValueError:
        return '0'

############# Moyenne 2 années H2H
def B(ratios):
    parsed = {}
    for item in ratios:
        for entry in item:
            name, value = entry.split(':')
            parsed[name.strip()] = float(value.strip())
    return parsed

def A(list1, list2):
    parsed1 = B(list1)
    parsed2 = B(list2)
    result = {}
    for key in set(parsed1.keys()).union(parsed2.keys()):
        value1 = parsed1.get(key, 0)
        value2 = parsed2.get(key, 0)
        result[key] = int((value1 + value2) / 2)
    return [[f"{key}: {value}"] for key, value in result.items()]

def Moy_H2H_2ans(list1, list2):
    final_result = []

    for item1, item2 in zip(list1, list2):
        if item1 == ['0', '0'] and item2 == ['0', '0']:
            final_result.append(['0', '0'])
        elif item1 == ['0', '0']:
            final_result.append(item2)
        elif item2 == ['0', '0']:
            final_result.append(item1)
        elif isinstance(item1, list) and all(isinstance(subitem, list) for subitem in item1) and isinstance(item2, list) and all(isinstance(subitem, list) for subitem in item2):
            result = A(item1, item2)
            final_result.append(result)
        else:
            final_result.append(['0', '0'])

    return final_result

#############
def extract_Tournois():
    home = 'https://www.tennisexplorer.com/matches/'
    racine = 'https://www.tennisexplorer.com'
    page = extract_source(home)

    if not page:
        print(f"No tables found for URL: {home}")
        return

    main_tournaments_section = page.find('td', string='Main tournaments').find_parent('tr').find_next_siblings('tr', class_='one')
    main_tournaments = []

    for row in main_tournaments_section:
        if row.find_previous_sibling('tr', class_='head').find('td').text.strip() != 'Main tournaments':
            break

        # Utiliser .get() pour récupérer les valeurs avec un fallback si non trouvé
        name_element = row.find('td', class_='t-name').find('a')
        name = name_element.text.strip() if name_element else " "

        link_element = row.find('td', class_='t-name').find('a')
        link = racine + link_element['href'] if link_element else " "

        surface_element = row.find('td', class_='s-color').find('span')
        surface = surface_element['title'] if surface_element else " "

        if row.find('span', class_='type-men2'):
            tournament_type = 'ATP'
        elif row.find('span', class_='type-women2'):
            tournament_type = 'WTA'
        else:
            tournament_type = 'Unknown'

        main_tournaments.append([tournament_type, name, link, surface])

    return main_tournaments


def fichematch_url(url):
    """Extraire les informations d'un match à partir de l'URL fournie."""
    #try:
    HTH = []
    page = extract_source(url)
    if not page:
        return None

    tournament_data = page.find('div', id='tournamentTabs-1-data')
    if not tournament_data:
        return None

    countrytournoi = re.search(r'\((.*?)\)', page.find('h1', class_='bg').text).group(1) if page.find('h1', class_='bg') else ""

    table = tournament_data.find('table', class_='result')
    if not table:
        return None

    tr_elements = table.find_all('tr')[1:]
    ficheurls, ficheh2h, fichedate, ficheheure, fichename, fichecote, ficheround = [], [], [], [], [], [], []

    aujourdhui = datetime.now()
    date_aujourdhui = aujourdhui.strftime("%d.%m.")
    date_demain = (aujourdhui + timedelta(days=1)).strftime("%d.%m.")

    for tr in tr_elements:
        time_td = tr.find('td', class_='first time')
        if time_td:
            date_heure = time_td.get_text(strip=True).replace('tomorrow', date_demain).replace('today', date_aujourdhui).split(', ')
            fichedate.append([date_heure[0]])
            ficheheure.append([date_heure[1]])

        round_td = tr.find('td', class_='round')
        if round_td:
            ficheround.append([round_td.get_text(strip=True)])

        a = tr.find('td', class_='t-name').find('a', href=True)
        if a and a['href'].startswith('/match-detail/'):
            ficheurls.append(f"https://www.tennisexplorer.com{a['href']}")

        h2h_cell = tr.find('td', class_='h2h')
        if h2h_cell:
            h2 = h2h_cell.get_text()
            ficheh2h.append([h2])
            HeadToHead = calculer_ratio(h2, '-')
            HTH.append([f'{math.ceil(HeadToHead[0])}', f'{math.ceil(HeadToHead[1])}'])

        t_name = tr.find('td', class_='t-name')
        if t_name:
            tnamef = re.sub(r'\(\d+\)', '', t_name.get_text()).strip()
            fichename.append([tnamef])

        t_cote = tr.find_all('td', class_='course')
        fichecote.extend(cote.get_text().replace(".", ",") for cote in t_cote)

    if not ficheurls or not ficheh2h or not fichename:
        print("Extraction incomplète : URLs ou données H2H manquants.")
        return None

    fichecote = paires(fichecote)
    fichecote = [[item.replace('\xa0', '') for item in sublist] for sublist in fichecote]
    ficheliengoogle = Lien_Google(fichename)

    fiche = [fichename, fichedate, ficheheure, ficheurls, fichecote, ficheh2h, ficheliengoogle, countrytournoi, ficheround]#, nom_J, class_J]

    fiche_opt = [
        [matchs[0] for matchs in fiche[0]],
        [date[0] for date in fiche[1]],
        [time[0] for time in fiche[2]],
        fiche[3],
        fiche[4],
        [result[0] for result in fiche[5]],
        [search[0] for search in fiche[6]],
        fiche[7],
        [round[0] for round in fiche[8]]
    ]

    return ficheurls, HTH, fiche, fiche_opt

    #except Exception as e:
        #print(f"Erreur: {e}")
        #return None

def create_match(year, tournament, surface, player_one, player_two, sets_one, sets_two, scores_one, scores_two):
    winner = player_one if sets_one > sets_two else player_two
    return {
        'Année': year,
        'Tournoi': tournament,
        'Rencontre': f"{player_one} - {player_two}",
        'Score_Final': f"{sets_one} - {sets_two}",
        'Gagnant': winner,
        'Surface': surface,
        'Joueur_1': player_one,
        'Jeu_gagnés_1': ', '.join(scores_one),
        'Sets_gagnés_1': sets_one,
        'Joueur_2': player_two,
        'Jeu_gagnés_2': ', '.join(scores_two),
        'Sets_gagnés_2': sets_two,
    }

def fichejoueur_url(urls):
    """Extraire les URLs des joueurs à partir des matchs."""
    liste, urls_img_J, nom_class, matches = [], [], [], []
    racine = 'https://www.tennisexplorer.com'

    for url in urls:
        match_tp = []
        verif = True
        page = extract_source(url)
        if not page:
            continue

        # Extraction des URLs des joueurs
        for result in page.find_all('th', class_='plName'):
            a = result.find('a', href=True)
            if a and a['href'].startswith('/player/'):
                liste.append(f"{racine}{a['href']}")

        # Extraction des images des joueurs
        for img_tag in page.find_all('td', class_='thumb'):
            img = img_tag.find('img', src=True)
            if img:
                urls_img_J.append(f"{racine}{img['src']}")

        # Extraction des tableaux de résultats
        tables = page.find_all('table', class_='result')
        try:
            table = tables[4]
            rows = table.find_all('tr', class_=['one', 'two'])
            for i in range(0, len(rows), 2):
                row_one = rows[i]
                row_two = rows[i + 1]
                columns_one = row_one.find_all('td')
                columns_two = row_two.find_all('td')

                year = columns_one[0].get_text()

            # Vérification si l'année est valide
                try:
                    year = int(year)
                except ValueError:
                    verif = False

                tournament = columns_one[1].get_text()
                span_element = columns_one[4].find('span')
                surface = span_element['title'] if span_element else "Info non disponible"
                player_one = columns_one[2].get_text()
                player_two = columns_two[0].get_text()
                sets_one = columns_one[3].get_text()
                sets_two = columns_two[1].get_text()
                scores_one = [col.get_text().strip() for col in columns_one[5:10]]
                scores_two = [col.get_text().strip() for col in columns_two[2:7]]

                match = create_match(year, tournament, surface, player_one, player_two, sets_one, sets_two, scores_one, scores_two)
                if verif:
                  match_tp.append(match)
                else:
                  match_tp.append('')

            df = pd.DataFrame(match_tp)
            matches.append(df)

        except IndexError:
            matches.append('')
            continue

    urls_J = [liste[i:i + 2] for i in range(0, len(liste), 2)]
    urls_img_J = [urls_img_J[i:i + 2] for i in range(0, len(urls_img_J), 2)]

    return urls_J, urls_img_J, matches

def extract_player_info(page):
    player_details = [page[0].find('h3').text]
    player_details.extend(detail.text for detail in page[0].find_all('div', class_='date'))
    return player_details

def create_match_tab(year, summary, surf_clay, surf_hard, surf_indoor, surf_grass, not_set):
    return {
        'Année': year,
        'Sommaire': summary,
        'Clay': surf_clay,
        'Hard': surf_hard,
        'Indoors': surf_indoor,
        'Grass': surf_grass,
        'Notset': not_set
    }

def process_rangs(rangs):
    rang = [
        [
            item.split('/')[0].strip().replace('.', '') if '/' in item else "-"
            for item in sublist
        ]
        for sublist in rangs
    ]

    rang_best = [
        [
            item.split('/')[1].strip().replace('.', '') if '/' in item and len(item.split('/')) > 1 else "-"
            for item in sublist
        ]
        for sublist in rangs
    ]

    return rang, rang_best

def Tableau(urls):
    career_J, info_J, player_names, match_J = [], [], [], []
    for url_list in urls:
        match_tp = []
        for url in url_list:
            if not url:
                match_tp.append(create_match_tab('', '', '', '', '', '', ''))
                continue
            page = extract_source(url)

############################################career_J, fiche_J#############################
            if not page:
                match_tp.append(create_match_tab('Er', 'Er', 'Er', 'Er', 'Er', 'Er', 'Er'))
                continue

            table_detail = page.find_all('table', class_='plDetail')
            info_J.append(extract_player_info(table_detail))

            tables = page.find_all('table', class_='result balance')
            if not tables:
                print(f"No tables found for URL: {url}")
                match_tp.append(create_match_tab('Er', 'Er', 'Er', 'Er', 'Er', 'Er', 'Er'))
                continue

            tr_with_photo = page.select_one('tr:has(td.photo)')
            player_name = tr_with_photo.find('h3').text if tr_with_photo else 'Unknown'
            player_names.append(player_name)


            table = tables[0]
            rows = table.find_all('tr', class_=['one', 'two'])
            for row in rows:
                columns = row.find_all('td')
                match_tp.append(create_match_tab(
                    columns[0].get_text(),  # year
                    columns[1].get_text(),  # summary
                    columns[2].get_text(),  # surf_clay
                    columns[3].get_text(),  # surf_hard
                    columns[4].get_text(),  # surf_indoor
                    columns[5].get_text(),  # surf_grass
                    columns[6].get_text()   # not_set
                ))

            if match_tp:
              career_J.append(pd.DataFrame(match_tp))
              match_tp = []

####### Sur le dernier mois nbr de match joués / défaite en favori / victoire en outsider
            match_J.append(process_matches(tables))


########################################################################################
    career_J = [career_J[i:i + 2] for i in range(0, len(career_J), 2) if i + 1 < len(career_J)]

    prefixes = ('Country:', 'Age:', 'Plays:','Current/Highest rank - singles:')

    def get_element(sub_list, prefix):
        return next((element for element in sub_list if element.startswith(prefix)), '')

    info_J = [[get_element(sub_list, prefix) for prefix in prefixes] for sub_list in info_J]

    info_J = [[item[0].split(': ')[1] if ': ' in item[0] else '',
                item[1].split(' (')[0].split(': ')[1] if ': ' in item[1] else '',
                item[2].split(': ')[1] if ': ' in item[2] else '',
                item[3].split(': ')[1] if ': ' in item[3] else ''
                ]
              for item in info_J]

    countries, ages, hands, rangs = [paires(list(x)) for x in zip(*info_J)]

    rang, rang_best = process_rangs(rangs)
    namefull_J = paires(player_names)

    fiche_J = countries, ages, hands, rang, rang_best, namefull_J

    return career_J, fiche_J,match_J

####
def Win_Car_Surf(urls, surf):
    PVCratio, PVSratio = [], []

    def traiter_page(page, surf):
        summary_tr = page.find('tr', class_='summary')
        if summary_tr:
            tds = summary_tr.find_all('td')
            get_text = lambda td: td.a.text if td.a else 'None'
            datas = [get_text(td) for td in tds[1:7]]
            surfaces = {"Summary": datas[0], "Clay": datas[1], "Hard": datas[2], "Indoors": datas[3], "Grass": datas[4], "Not set": datas[5]}
            dataS = surfaces.get(surf, "None")
            PVC = calculer_ratio(datas[0], '/')[0]
            PVCratio.append(f'{math.ceil(PVC)}')
            PVS = calculer_ratio(dataS, '/')[0]
            PVSratio.append(f'{math.ceil(PVS)}')
    for groupe_urls in urls:
        for url in groupe_urls:
            page = extract_source(url)
            if page:
                traiter_page(page, surf)
                tbody = page.find('table', class_='result balance')
                if tbody:
                    two_tr = tbody.find('tr', class_='two')
                    if two_tr:
                       tds = two_tr.find_all('td')
    return paires(PVCratio), paires(PVSratio)

########
def calcul_ratio_h2h(df):
    compteur = {}
    for _, row in df.iterrows():
        joueur_1, joueur_2, gagnant = row['Joueur_1'], row['Joueur_2'], row['Gagnant']
        if joueur_1 not in compteur:
            compteur[joueur_1] = {'victoires': 0, 'defaites': 0}
        if joueur_2 not in compteur:
            compteur[joueur_2] = {'victoires': 0, 'defaites': 0}
        if gagnant == joueur_1:
            compteur[joueur_1]['victoires'] += 1
            compteur[joueur_2]['defaites'] += 1
        else:
            compteur[joueur_2]['victoires'] += 1
            compteur[joueur_1]['defaites'] += 1
    resultat = []
    for joueur, stats in compteur.items():
        total_matches = stats['victoires'] + stats['defaites']
        if total_matches > 0:
            ratio = stats['victoires'] / total_matches * 100
        else:
            ratio = 0
        resultat.append([f"{joueur}: {'{:.0f}'.format(ratio)}"])
    return resultat


def h2h_an_surf(h2h_Tab, column, value, audit=False):
    filtered = []

    for ele in h2h_Tab:
        if isinstance(ele, pd.DataFrame) and column in ele.columns:
            ele['Année'] = ele['Année'].astype(str)
            filtered.append(ele.query(f"{column} == '{value}'"))
            if audit:
                print(f"\nh2h_Tab {column} {value}\n " + "#" * 20)
                print(ele.to_string())
        else:
            filtered.append('')

    result = [['0', '0'] if not isinstance(ele, pd.DataFrame) or ele.empty else calcul_ratio_h2h(ele) for ele in filtered]

    if audit:
        print(f"\nh2h_an_surf {column} {value}\n " + "#" * 20)
        print(result)

    return result

def sets_win(df):

    # print(len(df)," df avant:",df)
    # print(f"Nombre de lignes: {len(df)}, DataFrame avant traitement:")

    # Vérification combinée pour les cas invalides

    if (not isinstance(df, pd.DataFrame) or df.empty or
        df.isnull().all().all() or (df == 0).all().all() or
        df.replace('', float('nan')).isnull().all().all()):
        #print('###0,0###')
        return ['0', '0']

    #print(df.to_string())

    joueurs = df['Rencontre'].iloc[0].split(' - ')
    joueur_1, joueur_2 = joueurs[0], joueurs[1]
    sets_gagnes = {joueur_1: 0, joueur_2: 0}
    sets_joues = {joueur_1: 0, joueur_2: 0}
    for _, row in df.iterrows():
        sets_gagnes_1 = int(row['Sets_gagnés_1']) if row['Sets_gagnés_1'].isdigit() else 0
        sets_gagnes_2 = int(row['Sets_gagnés_2']) if row['Sets_gagnés_2'].isdigit() else 0
        sets_gagnes[row['Joueur_1']] += sets_gagnes_1
        sets_gagnes[row['Joueur_2']] += sets_gagnes_2
        sets_joues[row['Joueur_1']] += sets_gagnes_1 + sets_gagnes_2
        sets_joues[row['Joueur_2']] += sets_gagnes_1 + sets_gagnes_2
    pourcentage_1 = math.ceil(100 * sets_gagnes[joueur_1] / sets_joues[joueur_1]) if sets_joues[joueur_1] > 0 else 0
    pourcentage_2 = math.ceil(100 * sets_gagnes[joueur_2] / sets_joues[joueur_2]) if sets_joues[joueur_2] > 0 else 0

    if df['Année'].iloc[0]:
        resultat = [f"{joueur_1}: {pourcentage_1}", f"{joueur_2}: {pourcentage_2}"]
    else:
        resultat = ['0', '0']

    return resultat

def win_perso(career_J,surf,an, audit=False):
    a, b, e, f ,bb ,ff = [], [], [], [], [], []

    for htablist in career_J:
        for df in htablist:
            if not isinstance(df, pd.DataFrame):
                a.append('-')
                e.append('-')
            else:
                df_a = df[df['Année'] == f'{an}']
                #print(f"df {an}:\n", df_a.to_string())
                if not df_a.empty:
                    c = df_a.loc[df_a['Année'] == f'{an}', 'Sommaire'].values[0]
                    a.append(c)
                    d = df_a.loc[df_a['Année'] == f'{an}', f'{surf}'].values[0]
                    e.append(d)
                else :
                    a.append('-')
                    e.append('-')

    if audit:
        print("\nwin_perso\n" + "#" * 20)
        print(f"Sommaire (a):\n",len(a),a)
        print(f"Surf (e):\n",len(e),e)

      # Calculer les ratios en utilisant les compréhensions de liste
    for ele in a:
      b.append(f"{math.ceil(calculer_ratio(ele,'/')[0])}")
      bb.append(f"{math.ceil(calculer_ratio(ele,'/')[0])}")
    for ele in e:
      f.append(f"{math.ceil(calculer_ratio(ele,'/')[0])}")
      ff.append(f"{math.ceil(calculer_ratio(ele,'/')[0])}")

    return paires(b), paires(f)


def win_perso_moy_2ans(career_J, surf,lastan, audit=False):

    current_year = str(datetime.now().year)
    h_2023, i_2023 = win_perso(career_J, surf, lastan)
    h_2024, i_2024 = win_perso(career_J, surf, current_year)

    moyenne_h = []
    moyenne_i = []

    for ratio_2023, ratio_2024 in zip(h_2023, h_2024):
        ratio_moyen = [math.ceil((int(ratio_2023[0]) + int(ratio_2024[0])) / 2), math.ceil((int(ratio_2023[1]) + int(ratio_2024[1])) / 2)]
        moyenne_h.append(ratio_moyen)

    for ratio_2023, ratio_2024 in zip(i_2023, i_2024):
        ratio_moyen = [math.ceil((int(ratio_2023[0]) + int(ratio_2024[0])) / 2), math.ceil((int(ratio_2023[1]) + int(ratio_2024[1])) / 2)]
        moyenne_i.append(ratio_moyen)

    if audit:
        print("\nwin_perso_moy_2ans\n" + "#" * 20)
        print("Ratio Somm 23/24 (h):\n",h_2023, h_2024)
        print("Moy Somm (moy_h):\n",moyenne_h)
        print("Ratio Surf 23/24 (i):\n",i_2023, i_2024)
        print("Moy Surf (moy_i):\n",moyenne_i)

    return moyenne_h, moyenne_i

################ win_perso_10_50 #######################

def pourparan(data, audit = False):
    j, k = [], []
    total_matchs = 0
    nbrvict, nbrdef = 0, 0
    match10mini, match50mini = None, None
    nbrvict10, nbrdef10, nbrvict50, nbrdef50 = 0, 0, 0, 0

    if audit:
        print("\nwin_perso_10_50\n" + "#" * 20)

    for index, row in data.iterrows():
        try:
            # Check if 'Sommaire' exists and is valid
            if pd.isna(row['Sommaire']) or row['Sommaire'] == '' or '/' not in str(row['Sommaire']):
                if audit:
                    print(f"Skipping invalid 'Sommaire' value: {row['Sommaire']}")
                continue
            
            # Split and validate the format
            sommaire_parts = str(row['Sommaire']).split('/')
            if len(sommaire_parts) != 2:
                if audit:
                    print(f"Skipping invalid 'Sommaire' format: {row['Sommaire']}")
                continue
            
            # Try to convert to integers
            victoires, defaites = [int(v.strip()) for v in sommaire_parts]
            nbrvict += victoires
            nbrdef += defaites
            total_matchs += victoires + defaites
        except (ValueError, TypeError) as e:
            if audit:
                print(f"Error parsing 'Sommaire' value '{row['Sommaire']}': {e}")
            continue

        #if audit:
            #print(f"Index: {index}, Victoires: {victoires}, Defaites: {defaites}")
            #print(f"Total matchs: {total_matchs}, Nombre de victoires: {nbrvict}, Nombre de défaites: {nbrdef}")

        if total_matchs > 10 and match10mini is None:
            match10mini = total_matchs
            nbrvict10 = nbrvict
            nbrdef10 = nbrdef
            if audit: print(f"Matchs après 10: {match10mini}, Victoires: {nbrvict10}, Defaites: {nbrdef10}")

        if total_matchs > 50:
            match50mini = total_matchs
            nbrvict50 = nbrvict
            nbrdef50 = nbrdef
            if audit: print(f"Matchs après 50: {match50mini}, Victoires: {nbrvict50}, Defaites: {nbrdef50}")
            break
        elif total_matchs > 10:
            match50mini = total_matchs
            nbrvict50 = nbrvict
            nbrdef50 = nbrdef

    if match10mini is not None:
        ele10 = f"{math.ceil(calculer_ratio(f'{nbrvict10}/{nbrdef10}','/')[0])}"
        j = f'{ele10}'
        if audit: print(f"Ratio après 10 matchs: {ele10}")

    if match50mini is not None:
        ele50 = f"{math.ceil(calculer_ratio(f'{nbrvict50}/{nbrdef50}','/')[0])}"
        k = f'{ele50}'
        if audit: print(f"Ratio après 50 matchs: {ele50}")

    return j, k

def win_perso_10_50(career_J):
    j, k = [], []


    for n in range(len(career_J)):
        for df in career_J[n]:
            a, b = pourparan(df)
            j.append(a)
            k.append(b)
            #if audit: print(f"Résultats intermédiaires pour n={n}: j={j}, k={k}")

    j = paires(j)
    k = paires(k)

    return j, k

########################################################

def process_matches(table):
    def create_match(date, matches, resultat, cote_H, cote_A):
        return {
            'Date': date,
            'Match': matches,
            'Resultat': resultat,
            'CoteH': cote_H,
            'CoteA': cote_A,
        }

    def corriger_scores(scores):
        corrected_scores = []
        score_pattern = re.compile(r'(\d{1,2})-(\d{1,2})')
        for score in scores:
            corrected_parts = []
            parts = score.split(',')
            for part in parts:
                part = part.strip()
                match = score_pattern.match(part)
                if match:
                    set1, set2 = match.groups()
                    corrected_part = f'{set1[0]}-{set2[0]}'
                    corrected_parts.append(corrected_part)
            corrected_scores.append(', '.join(corrected_parts))
        return corrected_scores

    def extraire_matchs(data):
        dates, matchs, scores, cotes_H, cotes_A = [], [], [], [], []
        rows = data.find_all('tr', class_=['one', 'two'])
        if not rows:
            print("No matching rows found.")
            return [], [], [], [], [], []
        try:
            for row in rows:
                date_element = row.select_one('td.first.time')
                match_element = row.select_one('td.t-name')
                score_element = row.select_one('td.tl')
                cote_elements = row.select('td.course')
                if cote_elements:
                    cote_H_element = cote_elements[0]
                    cote_A_element = cote_elements[1]
                else:
                    cote_H_element = '1'
                    cote_A_element = '1'
                date = date_element.text if date_element else "-"
                match = match_element.text.strip().replace('\n', ' ') if match_element else "-"
                score = score_element.text if score_element else "-"
                cote_H = cote_H_element.text if cote_H_element else "-"
                cote_A = cote_A_element.text if cote_A_element else "-"
                dates.append(date)
                matchs.append(match)
                scores.append(score)
                cotes_H.append(cote_H)
                cotes_A.append(cote_A)
            corrected_scores = corriger_scores(scores)
        except Exception as e:
            return [], [], [], [], [], []  # Retourne des listes vides en cas d'erreur
        return dates, matchs, scores, corrected_scores, cotes_H, cotes_A

    def creer_df(dates, matchs, corrected_scores, cotes_H, cotes_A):
        data = []
        for i in range(len(dates)):
            match_data = create_match(dates[i], matchs[i], corrected_scores[i], cotes_H[i], cotes_A[i])
            data.append(match_data)
        return pd.DataFrame(data)

    def corriger_score(score):
        score_pattern = re.compile(r'(\d+)-(\d+)')
        parts = score.split(',')
        corrected_parts = []

        for part in parts:
            part = part.strip()
            matches = score_pattern.findall(part)
            for match in matches:
                set1, set2 = match
                if int(set1) > 9 or int(set2) > 9:
                    set1 = set1[:1]
                    set2 = set2[:1]
                corrected_parts.append(f'{set1}-{set2}')
        return ', '.join(corrected_parts)

    def trouver_joueur_principalold(df):
        joueurs = {}
        for match in df['Match']:
            joueur1, joueur2 = map(str.strip, match.split(' - '))
            joueurs[joueur1] = joueurs.get(joueur1, 0) + 1
            joueurs[joueur2] = joueurs.get(joueur2, 0) + 1
        return max(joueurs, key=joueurs.get)

    def trouver_joueur_principal(df):
        joueurs = {}
        for match in df['Match']:
            joueur_data = list(map(str.strip, match.split(' - ')))
            joueur1, joueur2 = joueur_data if len(joueur_data) == 2 else (None, None)
            if joueur1:
                joueurs[joueur1] = joueurs.get(joueur1, 0) + 1
            if joueur2:
                joueurs[joueur2] = joueurs.get(joueur2, 0) + 1
        if joueurs:
            return max(joueurs, key=joueurs.get)
        else:
            return "Aucun joueur trouvé"

    def ajouter_colonnes_resultats(df):
        principal = trouver_joueur_principal(df)
        df['defaites_favori'] = 0
        df['victoires_outsider'] = 0

        for index, row in df.iterrows():
            match = row['Match']
            cote1, cote2 = row['CoteH'], row['CoteA']
            score = row['Resultat']

            try:
                joueur1, joueur2 = map(str.strip, match.split(' - '))
                cote1, cote2 = float(cote1), float(cote2)
            except ValueError:
                #print(f"erreur split {match}/{cote1}-{cote2}")
                continue
            score = corriger_score(score)
            sets = score.split(', ')
            total_sets_joueur1 = sum(int(s.split('-')[0]) for s in sets if '-' in s)
            total_sets_joueur2 = sum(int(s.split('-')[1]) for s in sets if '-' in s)
            gagnant = joueur1 if total_sets_joueur1 > total_sets_joueur2 else joueur2
            if principal == joueur1:
                if cote1 < cote2 and gagnant != joueur1:
                    df.at[index, 'defaites_favori'] = 1
                elif cote1 > cote2 and gagnant == joueur1:
                    df.at[index, 'victoires_outsider'] = 1
            elif principal == joueur2:
                if cote2 < cote1 and gagnant != joueur2:
                    df.at[index, 'defaites_favori'] = 1
                elif cote2 > cote1 and gagnant == joueur2:
                    df.at[index, 'victoires_outsider'] = 1
        return df

    if len(table)>2:
      dates, matchs, scores, corrected_scores, cotes_H, cotes_A = extraire_matchs(table[2])
    else:
      dates, matchs, scores, corrected_scores, cotes_H, cotes_A = extraire_matchs(table[1])
    if not dates and len(table)>2:
        #dates, matchs, scores, corrected_scores, cotes_H, cotes_A = extraire_matchs(table[3])
        dates, matchs, scores, corrected_scores, cotes_H, cotes_A = extraire_matchs(table[2])
    if not dates:
        match_data = create_match('', '', '', '', '')
        df = pd.DataFrame([match_data])
    else:
        df = creer_df(dates, matchs, corrected_scores, cotes_H, cotes_A)
    df = ajouter_colonnes_resultats(df)
    return df

def matchjouemois(lf):
    def calculate_monthly_stats(df):
        if not isinstance(df, pd.DataFrame):
            raise TypeError("L'entrée doit être un DataFrame pandas.")
        current_month = datetime.now().strftime(".%m.")
        previous_month = (datetime.now() - timedelta(days=30)).strftime(".%m.")
        pattern_current_month = re.compile(r'\d{2}' + re.escape(current_month))
        pattern_previous_month = re.compile(r'\d{2}' + re.escape(previous_month))
        current_month_data = df[df['Date'].str.contains(pattern_current_month)]
        previous_month_data = df[df['Date'].str.contains(pattern_previous_month)]
        current_month_defaites_favori = current_month_data['defaites_favori'].sum()
        current_month_victoires_outsider = current_month_data['victoires_outsider'].sum()
        current_month_matches = current_month_data.shape[0]
        previous_month_defaites_favori = previous_month_data['defaites_favori'].sum()
        previous_month_victoires_outsider = previous_month_data['victoires_outsider'].sum()
        previous_month_matches = previous_month_data.shape[0]
        tot_month_defaites_favori = current_month_defaites_favori + previous_month_defaites_favori
        tot_month_victoires_outsider = current_month_victoires_outsider + previous_month_victoires_outsider
        tot_month_matches = current_month_matches + previous_month_matches
        return tot_month_matches, tot_month_defaites_favori, tot_month_victoires_outsider
    m, o, p = [], [], []
    for ele in lf:
        m0, o0, p0 = calculate_monthly_stats(ele)
        m.append(f'{m0}')
        o.append(f'{o0}')
        p.append(f'{p0}')
    return paires(m), paires(o), paires(p)

##########
########## Import tennisabstract
##########

from datetime import datetime
from itertools import permutations
import math
from bs4 import BeautifulSoup

def load_players_from_csv(filename):
    """Charge les joueurs existants à partir du fichier CSV."""
    joueurs_dict = {}
    try:
        with open(filename, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header
            for row in reader:
                if row:
                    joueurs_dict[normalize_string(row[0])] = row[1]
    except FileNotFoundError:
        return {}
    return joueurs_dict

def normalize_string(s):
    """Normalise une chaîne de caractères en supprimant les espaces, les caractères spéciaux et en convertissant en minuscules."""
    s = s.lower()
    s = re.sub(r'\s+', '', s)  # Supprime les espaces
    s = re.sub(r'\W+', '', s)  # Supprime les caractères spéciaux
    return s

def generate_name_permutations(full_name):
    """Génère toutes les permutations possibles des noms."""
    names = full_name.split()
    permutations_list = [''.join(p) for p in permutations(names)]
    return [normalize_string(p) for p in permutations_list]

def search_player(joueurs_dict, player_name):
    """Cherche le joueur par nom (y compris permutations) et retourne le lien s'il est trouvé."""
    name_permutations = generate_name_permutations(player_name)
    if audit1:
        print("\nplayer_name: ",player_name)
        print("name_permutations: ",name_permutations)

    for permuted_name in name_permutations:
        if permuted_name in joueurs_dict:
            return joueurs_dict[permuted_name]
    return None

def initialize_driver():
    """Initialise le driver Firefox avec les options nécessaires."""
    firefox_options = Options()
    firefox_options.add_argument("--headless")
    driver = webdriver.Firefox(options=firefox_options)
    return driver

def fetch_page_source(url, driver):
    """Récupère le contenu HTML de la page spécifiée."""
    try:
        # Vérifier si le driver est encore valide
        if driver.current_window_handle:
            driver.get(url)
            time.sleep(5)  # Attendre que la page se charge complètement
            return driver.page_source
        else:
            print("Driver window is not available")
            return None
    except Exception as e:
        print(f"Erreur lors de la récupération de la page {url}: {e}")
        return None

def parse_additional_info(page_source):
    """Analyse et extrait les informations supplémentaires du joueur à partir du contenu HTML."""
    soup = BeautifulSoup(page_source, 'html.parser')
    recent_results_section = soup.find('table', id='recent-results')
    recent_results = []
    if recent_results_section:
        matches = recent_results_section.find_all('tr')
        for match in matches:
            cells = match.find_all('td')
            if len(cells) > 1:
                recent_results.append({
                    "Date": cells[0].text.strip(),
                    "Tournament": cells[1].text.strip(),
                    "Surface": cells[2].text.strip(),
                    "Rd": cells[3].text.strip(),
                    "Rang": cells[4].text.strip() if len(cells) > 4 else '',
                    "vRk": cells[5].text.strip() if len(cells) > 5 else '',
                    "lien": cells[6].text.strip() if len(cells) > 6 else '',
                    "Score": cells[7].text.strip() if len(cells) > 7 else '',
                    "DR": cells[8].text.strip() if len(cells) > 8 else '',
                    "A%": cells[9].text.strip() if len(cells) > 9 else '',
                    "DF%": cells[10].text.strip() if len(cells) > 10 else '',
                    "1stin": cells[11].text.strip() if len(cells) > 11 else '',
                    "1st%": cells[12].text.strip() if len(cells) > 12 else '',
                    "2nd%": cells[13].text.strip() if len(cells) > 13 else '',
                    "BPVsd": cells[14].text.strip() if len(cells) > 14 else '',
                    "Time": cells[15].text.strip() if len(cells) > 15 else ''
                })
    return recent_results

from datetime import datetime, timedelta

def process_last_tournament(matches):
    """Calcule les données pour le dernier tournoi valide et le temps de jeu des 7 derniers jours."""

    matches_sorted = sorted(matches, key=lambda x: datetime.strptime(x["Date"], '%d-%b-%Y'), reverse=True)

    last_tournament = None
    total_time_minutes = 0
    total_time_last_7_days = 0  # Nouveau compteur pour le temps des 7 derniers jours
    sum_1st_percent = 0
    count_1st_percent = 0
    total_bp_won = 0
    total_bp_attempts = 0
    sum_1stin = 0
    count_1stin = 0

    # Obtenir la date d'aujourd'hui et la date des 7 derniers jours, en ignorant l'heure
    today = datetime.now().date()
    seven_days_ago = today - timedelta(days=7)

    # Fonction pour réinitialiser les statistiques
    def reset_statistics():
        nonlocal total_time_minutes, sum_1st_percent, count_1st_percent, total_bp_won, total_bp_attempts, sum_1stin, count_1stin
        total_time_minutes = 0
        sum_1st_percent = 0
        count_1st_percent = 0
        total_bp_won = 0
        total_bp_attempts = 0
        sum_1stin = 0
        count_1stin = 0

    # Fonction pour traiter un match et calculer les statistiques
    def process_match(match):
        nonlocal total_time_minutes, total_time_last_7_days, sum_1st_percent, count_1st_percent, total_bp_won, total_bp_attempts, sum_1stin, count_1stin

        # Conversion de la date du match avec le formalisme '11-Aug-2024'
        match_date = datetime.strptime(match["Date"], '%d-%b-%Y').date()

        if audit: print(f"Match date: {match_date}, Seven days ago: {seven_days_ago}")  # Debug

        # Traiter le temps pour tous les matchs du dernier tournoi
        time_str = match.get("Time")
        if time_str:  # Si le temps est valide
            time_parts = time_str.split(':')
            if len(time_parts) == 2:  # Vérifie que le format est correct
                hours = int(time_parts[0])
                minutes = int(time_parts[1])
                time_in_minutes = hours * 60 + minutes
                total_time_minutes += time_in_minutes

                # Ajouter le temps si le match est dans les 7 derniers jours
                if seven_days_ago <= match_date <= today:
                    if audit : print(f"Adding {time_in_minutes} minutes to the total for the last 7 days.")
                    total_time_last_7_days += time_in_minutes
                #else:
                    #print("Match is not in the last 7 days.")

        # Traitement des pourcentages et statistiques si les données sont présentes
        if match.get("1st%"):
            sum_1st_percent += float(match["1st%"].replace('%', '').replace(',', '.'))
            count_1st_percent += 1

        if match.get("BPVsd"):
            bp_won, bp_attempts = map(int, match["BPVsd"].split('/'))
            total_bp_won += bp_won
            total_bp_attempts += bp_attempts

        if match.get("1stin"):
            sum_1stin += float(match["1stin"].replace('%', '').replace(',', '.'))
            count_1stin += 1

    # Boucle sur les matchs triés pour trouver le dernier tournoi valide
    for match in matches_sorted:
        tournament_name = match["Tournament"]

        if last_tournament is None:
            last_tournament = tournament_name  # Identifie le premier tournoi
        elif tournament_name != last_tournament and total_time_minutes == 0:
            # Affiche un message de contrôle avant de passer au tournoi suivant
            if audit1 : print(f"Changement de tournoi : Passage de '{last_tournament}' à '{tournament_name}' car aucun temps valide n'a été trouvé.")

            # Si on change de tournoi et que le précédent n'avait pas de données valides, on passe au suivant
            last_tournament = tournament_name
            reset_statistics()

        if match["Tournament"] == last_tournament:
            process_match(match)

    # Si à la fin, aucune donnée n'est trouvée, retourner une réponse par défaut
    if total_time_minutes == 0:
        return ['Aucun tournoi trouvé', '-1', '-1', '-1', '-1', '-1']

    # Calcul des moyennes après avoir trouvé un tournoi valide
    avg_1st_percent = round((sum_1st_percent / count_1st_percent) if count_1st_percent > 0 else 0, 1)
    bp_percentage = round((total_bp_won / total_bp_attempts * 100) if total_bp_attempts > 0 else 0, 1)
    last_1stin = round((sum_1stin / count_1stin) if count_1stin > 0 else 0, 1)

    # Formatage des pourcentages avec une valeur par défaut si nécessaire
    def format_percentage(value, default="-1"):
        if isinstance(value, str):
            value = value.replace('%', '')
        if value and float(value) != 0:
            return str(math.ceil(float(value)))
        return default

    # Initialiser last_1stin avant de l'utiliser
    last_1stin = format_percentage(last_1stin)
    avg_1st_percent = format_percentage(avg_1st_percent)
    bp_percentage = format_percentage(bp_percentage)

    # Retourner les informations du dernier tournoi valide et le temps des 7 derniers jours
    return [
        last_tournament,
        last_1stin,
        avg_1st_percent,
        bp_percentage,
        str(total_time_minutes),
        str(total_time_last_7_days)  # Ajouter le temps passé à jouer dans les 7 derniers jours
    ]


def get_last_tournament_info(player_name, joueurs_dict):
    """Fonction principale pour obtenir les informations du dernier tournoi."""
    player_link = search_player(joueurs_dict, player_name.replace('-', ' '))

    if player_link:
        driver = None
        try:
            driver = initialize_driver()
            player_page_source = fetch_page_source(player_link, driver)
            
            if player_page_source:
                additional_info = parse_additional_info(player_page_source)
                last_tournament_info = process_last_tournament(additional_info)
                if audit:
                    print("additional_info: ",additional_info)
                    print("last_tournament_info: ",last_tournament_info)
                return last_tournament_info
            else:
                print(f"Impossible de récupérer les données pour {player_name}")
                return ['Données indisponibles', '-1', '-1', '-1', '-1','-1']
        except Exception as e:
            print(f"Erreur lors du traitement de {player_name}: {e}")
            return ['Erreur de traitement', '-1', '-1', '-1', '-1','-1']
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass  # Ignorer les erreurs de fermeture du driver
    else:
        return ['Aucun joueur trouvé', '-1', '-1', '-1', '-1','-1']


########
def abstract(f_NomPrenom):
    cube = []
    joueurs_dict = load_players_from_csv(csv_url)

    for pair in f_NomPrenom:
        for player_name_input in pair:
            last_tournament_info = get_last_tournament_info(player_name_input, joueurs_dict)
            last_tournament_info.insert(0, player_name_input)
            #print(f"Résultats pour {player_name_input} : {last_tournament_info}\n")
            cube.append(last_tournament_info)

    cube = paires(cube)
    if audit1: print("cube: ",cube)

        # Utilisation de zip pour extraire les informations
    players, tournaments, stat1, stat2, stat3, stat4, stat5 = zip(*[
        zip(*match) for match in cube
    ])

    # Convertir les tuples en listes pour un affichage plus facile
    players = [list(p) for p in players]
    tournaments = [list(t) for t in tournaments]
    stat1 = [list(s) for s in stat1]
    stat2 = [list(s) for s in stat2]
    stat3 = [list(s) for s in stat3]
    stat4 = [list(s) for s in stat4]
    stat5 = [list(s) for s in stat5]

    tortuegenial = players, tournaments, stat1, stat2, stat3, stat4, stat5

    return tortuegenial


################## Export ###################

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def safe_value(value):
    if isinstance(value, list) and not value:
        return ""  # Convertit les listes vides en chaînes vides
    return value

def exportxls(data, sheet_name='Sheet1', filename='Result_data_export.xlsx'):

    def split_J(name_pairs):
        return [name[0].split(' - ') for name in name_pairs]

    def flat(lst):
        lst_f = []
        for item in lst:
            if isinstance(item[0], list):  # si c'est une liste imbriquée
                flat_item = [item[0][0], item[1][0]]
            else:  # sinon c'est déjà une liste plate
                flat_item = item
            # Ajoutez la logique de transformation des pourcentages ici si nécessaire
            lst_f.append(flat_item)
        return lst_f

    # Fonction pour traiter les éléments h2h
    def process_h2h(h2h_item, a_player, b_player):
        if h2h_item == '0':
            return h2h_item
        for i in range(2):
            if h2h_item[i][0].startswith(a_player):
                h2h_item[0], h2h_item[1] = h2h_item[i], h2h_item[1 - i]
                break
            elif h2h_item[i][0].startswith(b_player):
                h2h_item[0], h2h_item[1] = h2h_item[1 - i], h2h_item[i]
                break
        for i in range(2):
            if h2h_item[i][0] != '0':
                h2h_item[i][0] = h2h_item[i][0].split(': ')[1]
        return h2h_item

    # Fonction pour traiter les sets
    def process_sets(sets_item, a_player, b_player):
        if sets_item == '0':
            return sets_item
        if sets_item[0].startswith(a_player):
            sets_item[0], sets_item[1] = sets_item[0], sets_item[1]
        elif sets_item[0].startswith(b_player):
            sets_item[0], sets_item[1] = sets_item[1], sets_item[0]
        if sets_item[0] != '0' and sets_item[1] != '0':
            sets_item[0] = sets_item[0].split(': ')[1]
            sets_item[1] = sets_item[1].split(': ')[1]
        return sets_item

    def transf(fiche):
        #a, b, c, d, e, f, g, h, i, j, k, m, o, p, fiche, linkImg, fiche_det = df
        #urls_M, h2h, fiche, urls_J, urls_img_J, h2h_Tab, career_J, fiche_J, PVC, PVS, urls_h2h, h2h_lastan_f, h2h_an_f,h2h_2ans_f, h2h_surf_f, h2h_sets_win, win_2ans, win_2ans_surf, win_10, win_50, match_J, M_mois, Def_fav_mois, Vict_out_mois, tournoi = df
        match_data = []

        # vide
        vide = [' ' for _ in range(len(fiche[7][3]))]
        #vide = paires(vide)
        # zero
        zero = ['0' for _ in range(len(fiche[7][3]))]
        #zero = paires(zero)
        # -1
        moinsun = ['-1' for _ in range(len(fiche[7][3]))]
        #moinsun = paires(moinsun)

        # H2H Ratio victoire/défaite Ens
        h2h_win = fiche[1]
        # H2H Ratio victoire/défaite sur 2 ans e
        h2h_win_an = fiche[13]
        # H2H Ratio victoire/défaite surface f
        h2h_win_surf = fiche[14]
        # H2H Pourcentage sets gagnés g
        h2h_win_set = fiche[15]
        # Pourcentage victoires sur 2ans
        win_an = [[str(item1), str(item2)] for item1, item2 in fiche[16]]
        # Pourcentage victoires surface sur 2ans
        win_surf_an = [[str(item1), str(item2)] for item1, item2 in fiche[17]]
        # Pourcentage victoires sur les 50 derniers matchs
        win_50 = fiche[19]
        # Pourcentage victoires sur les 10 derniers matchs
        win_10 = fiche[18]
        # Nombre de match joués (dernier mois)
        match_M = fiche[21]
        # Nombre défaite en favori (dernier mois)
        def_fav_M = fiche[22]
        # Nombre victoire en outsider (dernier mois)
        vict_out_M = fiche[23]
        # pays det1
        f_pays = fiche[7][0]
        # age det2
        f_age = fiche[7][1]
        # main det3
        f_main = fiche[7][2]
        # Tournoi
        #f_tourn = [f'{fiche[24]}' for _ in range(len(fiche[7][3])*2)]
        #f_tourn = paires(f_tourn)
        f_tourn = [f'{fiche[24]}' for _ in range(len(fiche[7][3]))]
        # Rencontre a
        f_rencontre = fiche[2][0]
        # Joueurs
        f_joueurs = split_J(f_rencontre)
        # classement
        f_classement = fiche[7][3]
        # Nom Prenom
        f_NomPrenom = fiche[7][5]
        # PVC
        f_PVC =  fiche[8]
        # PVS
        f_PVS =  fiche[9]
        # Date
        f_date = fiche[2][1]
        # Heure
        f_heure = fiche[2][2]
        # Côtes
        f_cote = fiche[2][4]
        # H2H
        f_h2h = fiche[2][5]
        # LieuTournoi
        f_lieutourn = ([f'{fiche[2][7]}' for _ in range(len(fiche[7][3]))])
        # Round
        f_round = fiche[2][8]
        # link Image
        f_linkImg = fiche[4]
        # link Google
        f_linkgoogle = fiche[2][6]
        # link TennisExplorer
        f_linkTennisExplorer = [[url] for url in fiche[0]]
        #f_total_time_minutes
        f_total_time_last_7_days = fiche[25][6]
        #f_last_1stin
        f_last_1stin = fiche[25][2]
        #f_avg_1st_percent
        f_avg_1st_percent = fiche[25][3]
        #f_bp_percentage
        f_bp_percentage = fiche[25][4]

        if audit:
            print("f_total_time_last_7_days:\n",f_total_time_last_7_days)
            print("f_last_1stin:\n",f_last_1stin)
            print("f_avg_1st_percent:\n",f_avg_1st_percent)
            print("f_bp_percentage:\n",f_bp_percentage)

########## Mets les noms à la bonnes place
        for indice, (f_rencontre_item, h2h_win_an_item, h2h_win_surf_item, h2h_win_set_item) in enumerate(zip(f_rencontre, h2h_win_an, h2h_win_surf, h2h_win_set)):
            try:
                a_player, b_player = [name.strip() for name in f_rencontre_item[0].split(' - ')]
            except ValueError as e:
                print(f"Error processing item at index {indice}: {f_rencontre_item}")
                continue
            h2h_win_an[indice] = process_h2h(h2h_win_an_item, a_player, b_player)
            h2h_win_surf[indice] = process_h2h(h2h_win_surf_item, a_player, b_player)
            h2h_win_set[indice] = process_sets(h2h_win_set_item, a_player, b_player)

        h2h_win_an = flat(h2h_win_an)
        h2h_win_surf = flat(h2h_win_surf)
###########################################
 # Définition des variables zippées
        if audit:
          test = f_joueurs, f_NomPrenom,f_classement, f_PVC, f_PVS, h2h_win, h2h_win_an, h2h_win_surf, h2h_win_set,win_an, win_surf_an, win_50, win_10, match_M, moinsun, def_fav_M, vict_out_M, vide, f_cote,f_tourn, f_date, f_heure, f_rencontre, f_h2h, f_linkImg, f_linkgoogle, f_linkTennisExplorer,f_pays, f_age, f_main, f_lieutourn, f_round
          listing(test)

        # Correction du zip sans redondance de noms de variables et avec correction de la décompaction
        zipped_items = zip(f_joueurs, f_classement,
                           f_PVC, f_PVS,
                           h2h_win, h2h_win_an, h2h_win_surf, h2h_win_set,
                           win_an, win_surf_an, win_50, win_10,
                           match_M, f_total_time_last_7_days, def_fav_M, vict_out_M,
                           f_last_1stin, f_avg_1st_percent, f_bp_percentage,
                           vide, vide,
                           f_cote, f_tourn, f_date, f_heure,
                           f_rencontre, f_h2h,
                           f_linkImg, f_linkgoogle, f_linkTennisExplorer,
                           f_pays, f_age, f_main,
                           f_lieutourn, f_round#,vide
                           )

        for indice, items in enumerate(zipped_items, start=1):
            # Décompression correcte des items sans redondance de noms
            (f_joueurs_item, f_classement_item,
             f_PVC_item, f_PVS_item,
             h2h_win_item, h2h_win_an_item, h2h_win_surf_item, h2h_win_set_item,
             win_an_item, win_surf_an_item, win_50_item, win_10_item,
             match_M_item, f_total_time_last_7_days_item, def_fav_M_item, vict_out_M_item,
             f_last_1stin_item, f_avg_1st_percent_item, f_bp_percentage_item,
             vide_item, vide_item,
             f_cote_item, f_tourn_item, f_date_item, f_heure_item,
             f_rencontre_item, f_h2h_item,
             f_linkImg_item, f_linkgoogle_item, f_linkTennisExplorer_item,
             f_pays_item, f_age_item, f_main_item,
             f_lieutourn_item, f_round_item,
             #vide_item
             ) = items

            # Création des données pour Joueur A
            match_joueurA = [
                'Match ' + str(indice), 'JoueurA',
                f_joueurs_item[0], f_classement_item[0],
                f_PVC_item[0], f_PVS_item[0],
                h2h_win_item[0], h2h_win_an_item[0], h2h_win_surf_item[0], h2h_win_set_item[0],
                win_an_item[0], win_surf_an_item[0], win_50_item[0], win_10_item[0],
                match_M_item[0], f_total_time_last_7_days_item[0], def_fav_M_item[0], vict_out_M_item[0],
                f_last_1stin_item[0], f_avg_1st_percent_item[0], f_bp_percentage_item[0],
                vide_item, vide_item,
                f_cote_item[0], f_tourn_item, f_date_item[0], f_heure_item[0],
                f_rencontre_item[0], f_h2h_item[0],
                f_linkImg_item[0], f_linkgoogle_item[0], f_linkTennisExplorer_item[0],
                f_pays_item[0], f_age_item[0], f_main_item[0],
                f_lieutourn_item, f_round_item[0],
                #vide_item
            ]

            # Création des données pour Joueur B
            match_joueurB = [
                'Match ' + str(indice), 'JoueurB',
                f_joueurs_item[1], f_classement_item[1],
                f_PVC_item[1], f_PVS_item[1],
                h2h_win_item[1], h2h_win_an_item[1], h2h_win_surf_item[1], h2h_win_set_item[1],
                win_an_item[1], win_surf_an_item[1], win_50_item[1], win_10_item[1],
                match_M_item[1], f_total_time_last_7_days_item[1], def_fav_M_item[1], vict_out_M_item[1],
                f_last_1stin_item[1], f_avg_1st_percent_item[1], f_bp_percentage_item[1],
                vide_item, vide_item,
                f_cote_item[1], f_tourn_item, f_date_item[0], f_heure_item[0],
                f_rencontre_item[0], f_h2h_item[0],
                f_linkImg_item[1], f_linkgoogle_item[0], f_linkTennisExplorer_item[0],
                f_pays_item[1], f_age_item[1], f_main_item[1],
                f_lieutourn_item, f_round_item[0],
                #vide_item
            ]

            # Ajout des données dans la liste finale
            match_data.append(match_joueurA)
            match_data.append(match_joueurB)

        if audit:
          listing(match_data)

        df = pd.DataFrame(match_data, columns=['Match', 'Tableau', 'Nom', 'Classement',
                                               'Pourcentage victoire sur la Carrière',
                                               'Pourcentage victoire sur la Surface',
                                               'H2H Ratio victoire/défaite ensemble carrière',
                                               'H2H Ratio victoire/défaite sur 1 an (2023)',
                                               'H2H Ratio victoire/défaite surface',
                                               'H2H Pourcentage sets gagnés',
                                               'Pourcentage victoire sur la Carrière (2023)',
                                               'Pourcentage victoire sur la Surface (2023)',
                                               'Pourcentage victoires sur les 50 derniers matchs',
                                               'Pourcentage victoires sur les 10 derniers matchs',
                                               'Nombre de match joués sur le dernier mois',
                                               'Durée cumulée des matchs des 7 derniers Jours',
                                               'Nombre défaite en favori (dernier mois)',
                                               'Nombre victoire en outsider (dernier mois)',
                                               '% 1er service dernier tournoi',
                                               '% de points gagnés sur le 1er service',
                                               '% de balle de break sauvées',
                                               '',
                                               '',
                                               'Côtes',
                                               'Tournoi',
                                               'Date',
                                               'Heure',
                                               'Rencontre',
                                               'H2H',
                                               'Lien Photo',
                                               'Lien Google',
                                               'Lien TennisExplorer',
                                               'Pays',
                                               'Age',
                                               'Main',
                                               'Lieu du tournoi',
                                               'Round',
                                               #'Dernier tournoi'
                                               ])

        df_transposed = df.T
        return df_transposed

    df_detail = transf(data)
    sheet_name = data[24]

    # Ouvrir le fichier Excel existant ou en créer un nouveau
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        # Supprimer la feuille par défaut de manière sécurisée
        if wb.active:
            wb.remove(wb.active)
    
    sheet_name_detail = f'{sheet_name}'

    # Vérifier si l'onglet existe déjà et le supprimer
    if sheet_name_detail in wb.sheetnames:
        wb.remove(wb[sheet_name_detail])
    
    # Supprimer la feuille temporaire si elle existe
    if "Temp" in wb.sheetnames:
        wb.remove(wb["Temp"])
    
    # Créer l'onglet pour ce tournoi
    ws_detail = wb.create_sheet(title=sheet_name_detail)

    # Écriture des nouvelles données dans l'onglet detail
    for r in dataframe_to_rows(df_detail, index=True, header=False):
        safe_row = [safe_value(cell) for cell in r]  # Nettoyer les valeurs avant d'écrire
        ws_detail.append(safe_row)

    # Sauvegarder le fichier Excel
    wb.save(filename)


################################################################################################################################################
def go(url, surf='Hard', lastan='2024', tournoi="Tournoi"):
    urls_M, h2h, fiche, fiche_opt = fichematch_url(url)

    if not urls_M:
        return None

    urls_J, urls_img_J, h2h_Tab = fichejoueur_url(urls_M)

    career_J, fiche_J, match_J = Tableau(urls_J)

    PVC, PVS = Win_Car_Surf(urls_J, surf)

    urls_h2h = ['' if dd == ['0', '0'] else uu for uu,dd in zip(urls_M,h2h)]

    h2h_lastan_f = h2h_an_surf(h2h_Tab, 'Année', lastan)

    h2h_an_f = h2h_an_surf(h2h_Tab, 'Année', str(int(lastan) + 1))

    h2h_2ans_f = Moy_H2H_2ans(h2h_lastan_f,h2h_an_f)

    h2h_surf_f = h2h_an_surf(h2h_Tab, "Surface", surf)

    h2h_sets_win = [sets_win(ele) for ele in h2h_Tab]

    win_2ans, win_2ans_surf = win_perso_moy_2ans(career_J, surf,lastan)

    win_10_50 = win_perso_10_50(career_J)

    win_10 = win_10_50[0]

    win_50 = win_10_50[1]

    M_mois, Def_fav_mois, Vict_out_mois = matchjouemois(match_J)

    f_abstract = abstract(fiche_J[5])
#######
    Krilin = urls_M, h2h, fiche, urls_J, urls_img_J, h2h_Tab, career_J, fiche_J, PVC, PVS, urls_h2h, h2h_lastan_f, h2h_an_f,h2h_2ans_f, h2h_surf_f, h2h_sets_win, win_2ans, win_2ans_surf, win_10, win_50, match_J, M_mois, Def_fav_mois, Vict_out_mois, tournoi,f_abstract

######## Export xlsx ########
    Chemin_fichier_xlsx = './data/Result_data_export.xlsx' ##########chemin fichier export à changer
   
    # Utiliser le nom du tournoi comme nom d'onglet
    sheet_name = tournoi.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')

    exportxls(Krilin, sheet_name, Chemin_fichier_xlsx)

    print("\n#" + "#" * 20)
    print(f"Traitement de {tournoi} ...")
    print(f"Nombre de matchs: {len(urls_M)}")

    return Krilin

csv_url = "./data/joueurs_ATP_WTA.csv" ##########chemin fichier joueur ATP WTA à changer

def clear_excel_file(filename='./data/Result_data_export.xlsx'):
    """Vide complètement le fichier Excel en créant un nouveau fichier vide."""
    try:
        wb = Workbook()
        # Garder au moins une feuille vide pour éviter l'erreur "At least one sheet must be visible"
        ws = wb.active
        ws.title = "Temp"
        wb.save(filename)
        print(f"Fichier Excel vidé : {filename}")
    except Exception as e:
        print(f"Erreur lors du vidage du fichier Excel : {e}")



def main():
    sangohan = []

    """Point d'entrée principal du script."""
    current_year = datetime.now().year
    lastan = str(current_year - 1)
    tournaments = extract_Tournois()

    if not tournaments:
        print("Aucun tournoi trouvé.")
        return

    # Vider le fichier Excel avant de commencer
    clear_excel_file()
    print("Fichier Excel vidé. Début du traitement automatique...")

    print("Voici la liste des tournois disponibles:")

    # Sélectionner automatiquement tous les tournois valides (surface non vide)
    selected_tournaments = []
    for i, tournament in enumerate(tournaments):
        if tournament[3] != ' ':
            print(f"{i + 1}. {tournament[1]} ({tournament[0]}) - {tournament[3]}")
            selected_tournaments.append(tournament)

    valid_surfaces = ['Indoors', 'Clay', 'Hard', 'Grass']

    for tournament in selected_tournaments:
        url, surface = tournament[2], tournament[3]
        if surface not in valid_surfaces:
            print(f"Surface non adaptée pour le tournoi {tournament[1]}: {surface}")
            continue

        print("\nEn cours ...")
        sangohan.append(go(url, surface, lastan, f"{tournament[1]} ({tournament[0]})"))

        if audit:
            print("Info Tournois:", url, surface, lastan, f"{tournament[1]} ({tournament[0]})")

    return sangohan

if __name__ == "__main__":
    if audit:
        print("#" * 20 + "\n     AUDIT\n" + "#" * 20 + "\n")
    sangoku = main()

    print("\n" + "#" * 20)
    print("fin")


